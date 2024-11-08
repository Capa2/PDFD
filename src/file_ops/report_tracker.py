import os
from openpyxl import Workbook, load_workbook
from pathlib import Path
from utils import config_manager
from utils.logging_setup import detailed_logger

settings = config_manager.load_settings()
output_file_path = settings['output'] / "report_status.xlsx"

def initialize_report_file() -> None:
    if not os.path.exists(output_file_path):
        detailed_logger.info(f"Creating new report file at {output_file_path}")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Report Status"
        sheet.append(["ID", "URL", "Alt. URL", "Status"])
        workbook.save(output_file_path)

def update_report_status(
        id_value: str, 
        url: str, 
        alt_url: str, 
        status: str
) -> None:
    try:
        if not os.path.exists(output_file_path):
            initialize_report_file()

        workbook = load_workbook(output_file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == id_value:
                row[3].value = status
                workbook.save(output_file_path)
                detailed_logger.info(f"Updated report status for ID {id_value} to '{status}'")
                return

        sheet.append([id_value, url, alt_url, status])
        workbook.save(output_file_path)
        detailed_logger.info(f"Created report status for ID {id_value} as '{status}'")
    except Exception as e:
        detailed_logger.error(f"Failed to update report status for ID {id_value}: {e}")

def is_report_downloaded(
        report_id: str, 
        output_dir_path: str=settings['output']
) -> bool:
    output_dir = Path(output_dir_path)
    matching_files = list(output_dir.glob(f"{report_id}.*"))
    return len(matching_files) > 0

def get_report_status(id_value: str) -> str:
    try:
        if not Path(output_file_path).exists():
            detailed_logger.warning(f"Report status file not found at {output_file_path}")
            return "Unknown"

        workbook = load_workbook(output_file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == id_value:
                return row[3]

        return "Unknown"
    except Exception as e:
        detailed_logger.error(f"Failed to get report status for ID {id_value}: {e}")
        return "Error"